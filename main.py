# 百度通用翻译API,不包含词典、tts语音合成等资源，如有相关需求请联系translate_api@baidu.com
# coding=utf-8

import http.client
import hashlib
import urllib
import random
import json


import time
import os 
import xlwt
import xlrd

import openpyxl


def baidu_trans_api(q = 'dread'):
    """
    input: string to translte
    output: respond
    http://api.fanyi.baidu.com/
    """
    appid = '20180902000201888'  # 填写你的appid
    secretKey = 'dveMma2P96LBqJqQ6nar'  # 填写你的密钥
    httpClient = None
    myurl = '/api/trans/vip/translate'

    fromLang = 'auto'   #原文语种
    toLang = 'zh'   #译文语种
    salt = random.randint(32768, 65536)
    sign = appid + q + str(salt) + secretKey
    sign = hashlib.md5(sign.encode()).hexdigest()
    myurl = myurl + '?appid=' + appid + '&q=' + urllib.parse.quote(q) + '&from=' + fromLang + '&to=' + toLang + '&salt=' + str(
    salt) + '&sign=' + sign

    try:
        httpClient = http.client.HTTPConnection('api.fanyi.baidu.com')
        httpClient.request('GET', myurl)

        # response是HTTPResponse对象
        response = httpClient.getresponse()
        result_all = response.read().decode("utf-8")
        result = json.loads(result_all)

        # print (result)

    except Exception as e:
        print (e)
    finally:
        if httpClient:
            httpClient.close()

    if "trans_result" in result:
        return result["trans_result"][0]["dst"]
    else:
        return "Invalid Access Limit"


def get_files_path(root_dir="./CET6_txt/"):
    files_path = []
    for year in os.listdir(root_dir):
        root_dir_ = os.path.join(root_dir, year)
        for file in os.listdir(root_dir_):
            if file.split(".")[-1]  == "txt":
                files_path.append(os.path.join(root_dir_, file))
    # print(files_path)
    
    return files_path



def string2words_nums(doc):
    hashmap = {}
    # print(doc.split())
    for word in doc.lower().replace(",", " ").replace("\"", " ").replace("_", " ").replace(".", " ").replace("(", " ").replace(")", " ").replace("?", " ").split():

        if word not in hashmap:
            hashmap[word] = 1
        else: 
            hashmap[word] += 1
    hashmap_sorted= sorted(hashmap.items(), key=lambda d:d[1], reverse = True)
    return hashmap_sorted


def trans2chinese_save(hashmap_sorted):
    filename = 'resutls.xlsx'  # 保存文件名称
    workbook = openpyxl.Workbook(write_only=True)  # 创建一个工作簿  encoding='utf8', style_compression=0
    workbook.create_sheet(title="cet6_sheet")  # 创建一个sheet页
    workbook.save(filename)

    row_used = 0  # 统计单词的数量，当做单词的编号
    print("\nNumber of all words :", len(hashmap_sorted))
    for word_frequency in hashmap_sorted:
        word = word_frequency[0]
        freq = word_frequency[1]

        word_to_trans = word
        if len(word_to_trans) > 2:
            if freq <= 1000000:
                time.sleep(1)  # 百度翻译API的FPS（每秒请求数）为1，这里需要设置程序延时。
                dst = baidu_trans_api(q = word_to_trans)
                print("Translation results :", row_used, freq, "-->", word,":", dst)
                wb = openpyxl.load_workbook(filename=filename)
                work_sheet = wb['cet6_sheet']
                row_used += 1

                if row_used == 1:
                    work_sheet.cell(column=1, row=row_used, value="编号")  #  存放 单词的编号
                    work_sheet.cell(column=2, row=row_used, value="频率")  #　存放　频数
                    work_sheet.cell(column=3, row=row_used, value="待翻译单词")  # 存放 单词
                    work_sheet.cell(column=4, row=row_used, value="翻译结果")  # 存放 翻译后的中文
                work_sheet.cell(column=1, row=row_used+1, value=row_used)  #  存放 单词的编号
                work_sheet.cell(column=2, row=row_used+1, value=freq)  #　存放　频数
                work_sheet.cell(column=3, row=row_used+1, value=word)  # 存放 单词
                work_sheet.cell(column=4, row=row_used+1, value=dst)  # 存放 翻译后的中文
                
                
                wb.save(filename = filename)
       

 
        
# print(hashmap_sorted)




if __name__  ==  "__main__":
    
    root_dir="./CET6_txt/"
    paths = get_files_path(root_dir=root_dir)
    string_add = ""
    for path in paths:
        f = open(path)
        string = f.read()
        print("The number of this path‘s words is:", len(string), path)
        string_add += string
        f.close()
    hashmap_sorted = string2words_nums(string_add)  
    trans2chinese_save(hashmap_sorted)




